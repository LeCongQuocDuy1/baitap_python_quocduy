import openpyxl
import mysql.connector
from student import Student
listSV = []
mini = [3, 10]
gioi = 0
kha = 0
tb = 0


def getlistSV():
    return listSV

def SortCri(e):
    return e.MaSV


def tinhdtb(a, b, c):
    return round((a + b + c) / 3, 1)


def checkHK(e):
    if e >= 8:
        return "Giỏi"
    elif e >= 6.5:
        return "Khá"
    else:
        return "Trung Bình"


def max(i):
    max = 0
    if i == 0:
        for j in range(len(listSV)):
            if max < len(listSV[j].MaSV):
                max = len(listSV[j].MaSV)

    if i == 1:
        for j in range(len(listSV)):
            if max < len(listSV[j].Ho + " " + listSV[j].Ten):
                max = len(listSV[j].Ho + " " + listSV[j].Ten)

    if max < mini[i]:
        max = mini[i]

    return max


def convertdate(time):
    date= time.split("/")
    return date[2]+"-"+date[1]+"-"+date[0]


def readexcel():
    global hk
    global filename
    hk = [0, 0, 0]
    filename = input("Nhập tên file cần đọc:") + '.xlsx'
    book = openpyxl.load_workbook(filename)
    sh = book.active
    maxrow = sh.max_row
    i = 1
    t = "A" + str(i)
    while sh[t].value != "STT" or i == maxrow:
        i += 1
        t = "A" + str(i)
    if i == maxrow:
        print("Không đọc dược dữ liệu")
    else:
        global index
        index = i
        j = i + 1
        t = "A" + str(j)
        while type(sh[t].value) == int:

            stu = Student(sh.cell(row=j, column=2).value,
                          sh.cell(row=j, column=3).value,
                          sh.cell(row=j, column=4).value,
                          sh.cell(row=j, column=5).value,
                          round(float(sh.cell(row=j, column=6).value), 1),
                          round(float(sh.cell(row=j, column=7).value), 1),
                          round(float(sh.cell(row=j, column=8).value), 1),
                          0, ""
                          )
            stu.DTB = tinhdtb(stu.Toan, stu.Ly, stu.Hoa)
            stu.HK = checkHK(stu.DTB)
            if stu.HK == "Giỏi":
                hk[0] += 1
            elif stu.HK == "Khá":
                hk[1] += 1
            else:
                hk[2] += 1

            listSV.append(stu)
            j += 1
            t = "A" + str(j)
        listSV.sort(key=SortCri)


def showsv(list):
    print(len(list))
    totallen = 0
    for i in range(2):
        totallen += max(i)

    print("_" * (totallen + 39))
    print(
        "| {} | {} | {:^5} | {:^5} | {:^5} | {:^5} |".format("STT".center(max(0)), "Họ và Tên".center(max(1)),
                                                             "Toán", "Lý",
                                                             "Hóa", "ĐTB", "HK"))
    for i in list:
        print(
            "| {:} | {:} | {:^5} | {:^5} | {:^5} | {:^5} |".format(i.MaSV.center(max(0)),
                                                                   (i.Ho + " " + i.Ten).center(max(1)),
                                                                   i.Toan, i.Ly, i.Hoa,
                                                                   i.DTB, i.HK))
    print("|" + "_" * (totallen + 37) + "|")


def sqlconn():
    global myconn
    global cur
    try:

        myconn = mysql.connector.connect(host="localhost", user="root", passwd="", database="pythondb")
        cur = myconn.cursor()
    except mysql.connector.errors.ProgrammingError:
        myconn = mysql.connector.connect(host="localhost", user="root", passwd="")
        cur = myconn.cursor()
        cur.execute("CREATE DATABASE pythondb")


def insert(stu):
    sql = "INSERT INTO Students(masv, ho, ten, ngaysinh, toan, ly, hoa, dtb, hk)" \
          + "values(%s,%s,%s,%s,%s,%s,%s,%s,%s)"
    val = (stu.MaSV, stu.Ho, stu.Ten, convertdate(stu.NgaySinh), stu.Toan, stu.Ly, stu.Hoa, stu.DTB, stu.HK)
    try:
        cur.execute(sql, val)
        myconn.commit()
    except:
        myconn.rollback()
        print(cur.rowcount, "record inserted")
        myconn.close()


def insertall(list):
    for i in list:
        insert(i)